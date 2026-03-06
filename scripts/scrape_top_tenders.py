"""
Web scraper voor de top-scoring tenders uit de AI analyse.

Doel: voor elke top tender de contracteinddatum ophalen zodat Sales weet
      wanneer de tender opnieuw gepubliceerd wordt en tijdig actie kan ondernemen.

Wat doet dit script?
1. Laadt de TenderNed dataset + AI score cache
2. Selecteert de top N tenders op AI score (standaard top 100)
3. Voor tenders met bekende contract_end → direct gebruiken
4. Voor tenders zonder contract_end → TenderNed pagina scrapen
5. Berekent actiedatum Sales = contract_end - 4 maanden
6. Exporteert naar Excel (deelbaar met Sales)

Gebruik:
    python scripts/scrape_top_tenders.py
    python scripts/scrape_top_tenders.py --top 50
    python scripts/scrape_top_tenders.py --top 100 --min-score 60

Vereisten:
    pip install requests beautifulsoup4 openpyxl pandas
"""

import argparse
import hashlib
import json
import re
import sys
import time
from datetime import datetime
from pathlib import Path

import pandas as pd

# Windows console UTF-8
if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.config import LOCAL_DATASET_PATTERN, DEFAULT_LEAD_MONTHS
from src.data_loader import normalize_column_names, parse_dates, clean_text_columns

OUTPUT_DIR = Path(__file__).parent / "output"
CACHE_FILE = PROJECT_ROOT / "cache" / "ai_scores.json"
SCRAPE_CACHE_FILE = PROJECT_ROOT / "cache" / "scraped_contracts.json"

# Kolommen die duidelijk aangeven dat ze gescraped zijn (niet uit bronbestand)
SCRAPED_COLUMN_PREFIX = "scraped_"

DUTCH_MONTHS = {
    "jan": "01", "feb": "02", "mrt": "03", "apr": "04",
    "mei": "05", "jun": "06", "jul": "07", "aug": "08",
    "sep": "09", "okt": "10", "nov": "11", "dec": "12",
}


# ─── Cache helpers ────────────────────────────────────────────────────────────

def load_scrape_cache() -> dict:
    """Laad eerder gescrapete contractdata."""
    if SCRAPE_CACHE_FILE.exists():
        try:
            with open(SCRAPE_CACHE_FILE, encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            return {}
    return {}


def save_scrape_cache(cache: dict) -> None:
    SCRAPE_CACHE_FILE.parent.mkdir(exist_ok=True)
    with open(SCRAPE_CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


def load_ai_cache() -> dict:
    """Laad AI scores uit cache bestand."""
    if CACHE_FILE.exists():
        try:
            with open(CACHE_FILE, encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            return {}
    return {}


def make_tender_hash(title: str, description: str, organization: str) -> str:
    content = "|".join([str(title), str(description), str(organization)])
    return hashlib.md5(content.encode()).hexdigest()


# ─── Dataset laden ────────────────────────────────────────────────────────────

def find_dataset() -> Path | None:
    matches = sorted(
        PROJECT_ROOT.glob(LOCAL_DATASET_PATTERN),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return matches[0] if matches else None


def load_dataset(path: Path) -> pd.DataFrame | None:
    print(f"Dataset laden: {path.name}...")
    try:
        xlsx = pd.ExcelFile(path, engine="openpyxl")
        data_sheet = next(
            (s for s in xlsx.sheet_names if "opendata" in s.lower() or
             ("data" in s.lower() and "leeswijzer" not in s.lower())),
            xlsx.sheet_names[0]
        )
        df = pd.read_excel(path, sheet_name=data_sheet, engine="openpyxl")
        df = normalize_column_names(df)
        df = parse_dates(df)
        df = clean_text_columns(df)
        print(f"  {len(df):,} rijen geladen")
        return df
    except Exception as e:
        print(f"Fout bij laden dataset: {e}")
        return None


# ─── AI scores koppelen ───────────────────────────────────────────────────────

def merge_ai_scores(df: pd.DataFrame, ai_cache: dict) -> pd.DataFrame:
    """Voeg AI scores toe aan het DataFrame via hash-matching."""
    df = df.copy()
    df["ai_score"] = None
    df["ai_explanation"] = None
    df["ai_product"] = None
    df["ai_sector"] = None
    df["ai_confidence"] = None

    matched = 0
    for idx, row in df.iterrows():
        h = make_tender_hash(
            row.get("title", ""),
            row.get("description", ""),
            row.get("organization", ""),
        )
        if h in ai_cache and not ai_cache[h].get("error"):
            cached = ai_cache[h]
            df.at[idx, "ai_score"] = cached.get("relevance_score")
            df.at[idx, "ai_explanation"] = cached.get("explanation")
            df.at[idx, "ai_product"] = cached.get("best_product")
            df.at[idx, "ai_sector"] = cached.get("sector_match")
            df.at[idx, "ai_confidence"] = cached.get("confidence")
            matched += 1

    print(f"  AI scores gekoppeld: {matched:,} tenders")
    return df


# ─── TenderNed scraper ────────────────────────────────────────────────────────

def _parse_dutch_date(date_str: str) -> str | None:
    """
    Parst Dutch date formats naar ISO (YYYY-MM-DD):
    - "27 jan. 2022" → "2022-01-27"
    - "01-01-2022"   → "2022-01-01"
    - "2022-01-01"   → "2022-01-01"
    """
    date_str = date_str.strip().rstrip(".")

    # Dutch long format: "27 jan. 2022" of "27 januari 2022"
    m = re.match(r"(\d{1,2})\s+([a-z]+)\.?\s+(\d{4})", date_str, re.IGNORECASE)
    if m:
        day, month_str, year = m.groups()
        month_key = month_str[:3].lower()
        month = DUTCH_MONTHS.get(month_key)
        if month:
            return f"{year}-{month}-{int(day):02d}"

    # Numeric formats
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(date_str, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue

    return None


def scrape_tenderned_page(url: str, playwright_browser=None) -> dict:
    """
    Scrape een TenderNed aankondigingspagina met Playwright (JS-rendered).

    Haalt op:
    - contract_start / contract_end ("Aanvang opdracht" / "Voltooiing opdracht")
    - contract_duration (looptijd uit de omschrijving)
    - contract_value (definitieve waarde)
    - num_bids (aantal inschrijvers)

    Returns dict met gevonden velden.
    """
    result = {
        "scraped_contract_start": None,
        "scraped_contract_end": None,
        "scraped_contract_duration": None,
        "scraped_contract_value": None,
        "scraped_num_bids": None,
        "scraped_url": url,
        "scraped_at": datetime.now().strftime("%Y-%m-%d"),
        "scrape_success": False,
    }

    try:
        page = playwright_browser.new_page()
        page.goto(url, wait_until="networkidle", timeout=20000)
        text = page.inner_text("body")
        page.close()

        result["scrape_success"] = True

        def find_value_after_label(text: str, label: str, chars: int = 40) -> str | None:
            """Zoek de waarde direct na een label in de paginatekst."""
            idx = text.lower().find(label.lower())
            if idx == -1:
                return None
            return text[idx + len(label): idx + len(label) + chars].strip()

        # Datumpatronen (NL-formaat of ISO)
        date_re = re.compile(
            r"\b(\d{1,2}\s+[a-z]{3}\.?\s+\d{4}|\d{1,2}[-/]\d{1,2}[-/]\d{4}|\d{4}[-/]\d{1,2}[-/]\d{1,2})\b",
            re.IGNORECASE,
        )

        def find_date_after_label(text: str, labels: list[str]) -> str | None:
            for label in labels:
                snippet = find_value_after_label(text, label, 60)
                if not snippet:
                    continue
                if snippet.startswith("-"):
                    continue  # TenderNed toont "-" als de datum ontbreekt
                m = date_re.search(snippet)
                if m:
                    return _parse_dutch_date(m.group(1))
            return None

        # Contract startdatum
        result["scraped_contract_start"] = find_date_after_label(text, [
            "Aanvang opdracht\n",
            "Aanvang van de opdracht\n",
            "Ingangsdatum\n",
        ])

        # Contract einddatum
        result["scraped_contract_end"] = find_date_after_label(text, [
            "Voltooiing opdracht\n",
            "Voltooiing van de opdracht\n",
            "Einddatum\n",
        ])

        # Looptijd uit de omschrijving (als einddatum ontbreekt)
        if not result["scraped_contract_end"]:
            duration_match = re.search(
                r"(\d+)\s*(?:jaar|maanden?)",
                text,
                re.IGNORECASE,
            )
            if duration_match:
                result["scraped_contract_duration"] = duration_match.group(0)

        # Aantal inschrijvingen
        bids_match = re.search(
            r"(?:aantal\s+(?:ontvangen\s+)?inschrijvingen?)[^\d]*(\d+)",
            text,
            re.IGNORECASE,
        )
        if bids_match:
            result["scraped_num_bids"] = int(bids_match.group(1))

        # Contractwaarde
        value_match = re.search(
            r"(?:totale\s+waarde|definitieve\s+waarde|waarde\s+van\s+de\s+opdracht)[^\d€]*€?\s*([\d\s.,]+)",
            text,
            re.IGNORECASE,
        )
        if value_match:
            raw = value_match.group(1).strip().replace(" ", "").replace(".", "").replace(",", ".")
            try:
                result["scraped_contract_value"] = float(raw)
            except ValueError:
                pass

    except Exception as e:
        result["scrape_error"] = str(e)

    return result


# ─── Actiedatum berekenen ─────────────────────────────────────────────────────

def calculate_action_date(contract_end: str | None, lead_months: int = DEFAULT_LEAD_MONTHS) -> str | None:
    """
    Actiedatum voor Sales = contract_end - lead_months maanden.
    SkillsTown-norm: 4 maanden voor verwachte herpublicatie contact opnemen.
    """
    if not contract_end:
        return None
    try:
        end_dt = pd.to_datetime(contract_end)
        action_dt = end_dt - pd.DateOffset(months=lead_months)
        return action_dt.strftime("%Y-%m-%d")
    except Exception:
        return None


def days_until_action(action_date: str | None) -> int | None:
    if not action_date:
        return None
    try:
        delta = pd.to_datetime(action_date) - datetime.now()
        return delta.days
    except Exception:
        return None


def urgency_label(days: int | None) -> str:
    if days is None:
        return "Onbekend"
    if days < 0:
        return "VERLOPEN"
    if days <= 30:
        return "URGENT"
    if days <= 90:
        return "Hoog"
    if days <= 180:
        return "Medium"
    return "Laag"


# ─── Exporteer naar Excel ─────────────────────────────────────────────────────

def export_to_excel(df: pd.DataFrame, path: Path) -> None:
    """
    Exporteer naar een deelbare Excel met duidelijke kolommen.
    Gescrapete kolommen staan rechts en zijn duidelijk gelabeld.
    """
    # Kies welke kolommen we exporteren en in welke volgorde
    basis_cols = [
        "ai_score", "ai_explanation", "ai_product", "ai_sector", "ai_confidence",
        "title", "organization", "organization_city",
        "publication_date", "award_date",
        "contract_start", "contract_end",       # Uit bronbestand
        "contract_value", "winning_company",
        "cpv_codes", "tender_url",
    ]

    scraped_cols = [
        "scraped_contract_start",   # Gescraped van TenderNed
        "scraped_contract_end",
        "scraped_contract_duration",
        "scraped_contract_value",
        "scraped_num_bids",
        "contract_end_final",       # Beste beschikbare einddatum
        "contract_end_bron",        # Hoe betrouwbaar is de einddatum?
        "sales_action_date",        # Actiedatum voor Sales
        "days_until_action",        # Dagen tot actie
        "urgency",                  # URGENT / Hoog / Medium / Laag
        "scraped_url",
        "scraped_at",
        "scrape_success",
    ]

    export_cols = [c for c in basis_cols + scraped_cols if c in df.columns]
    export_df = df[export_cols].copy()

    # Mooiere kolomnamen voor Sales
    rename_display = {
        "ai_score": "AI Score (0-100)",
        "ai_explanation": "AI Toelichting",
        "ai_product": "SkillsTown Product",
        "ai_sector": "Sector",
        "ai_confidence": "AI Zekerheid",
        "title": "Tender Titel",
        "organization": "Organisatie",
        "organization_city": "Stad",
        "publication_date": "Publicatiedatum",
        "award_date": "Gunningsdatum",
        "contract_start": "Contractstart (bron)",
        "contract_end": "Contracteinde (bron)",
        "contract_value": "Contractwaarde (bron)",
        "winning_company": "Winnende Partij",
        "cpv_codes": "CPV Codes",
        "tender_url": "TenderNed Link",
        "scraped_contract_start": "Contractstart [gescraped]",
        "scraped_contract_end": "Contracteinde [gescraped]",
        "scraped_contract_duration": "Looptijd [gescraped]",
        "scraped_contract_value": "Contractwaarde [gescraped]",
        "scraped_num_bids": "Aantal Inschrijvers [gescraped]",
        "contract_end_final": "Contracteinde (beste schatting)",
        "contract_end_bron": "Betrouwbaarheid datum",
        "sales_action_date": "ACTIEDATUM SALES",
        "days_until_action": "Dagen tot Actie",
        "urgency": "Urgentie",
        "scraped_url": "Gescrapete URL",
        "scraped_at": "Gescraped op",
        "scrape_success": "Scrape Geslaagd",
    }
    export_df = export_df.rename(columns={k: v for k, v in rename_display.items() if k in export_df.columns})

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name="Top Tenders - Sales", index=False)

        # Opmaak
        ws = writer.sheets["Top Tenders - Sales"]
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        # Header rij opmaken
        header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        scraped_fill = PatternFill(start_color="1a3a5c", end_color="1a3a5c", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        scraped_display_cols = {rename_display.get(k, k) for k in [
            "scraped_contract_start", "scraped_contract_end", "scraped_contract_duration",
            "scraped_contract_value", "scraped_num_bids", "scraped_url", "scraped_at", "scrape_success"
        ]}

        for cell in ws[1]:
            if cell.value in scraped_display_cols:
                cell.fill = scraped_fill
            else:
                cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Urgentie kleuren
        urgency_col = None
        actie_col = None
        for i, cell in enumerate(ws[1], 1):
            if cell.value == "Urgentie":
                urgency_col = i
            if cell.value == "ACTIEDATUM SALES":
                actie_col = i

        urgency_colors = {
            "VERLOPEN": "8B0000",
            "URGENT": "FF4B4B",
            "Hoog": "FFA500",
            "Medium": "FFD700",
            "Laag": "90EE90",
            "Onbekend": "CCCCCC",
        }

        if urgency_col:
            for row in ws.iter_rows(min_row=2, min_col=urgency_col, max_col=urgency_col):
                for cell in row:
                    val = str(cell.value or "")
                    color = urgency_colors.get(val, "FFFFFF")
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    cell.font = Font(bold=(val in ("URGENT", "VERLOPEN")))

        # Kolombreedte aanpassen
        for col_idx, col in enumerate(ws.columns, 1):
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

        # Rijen afwisselend inkleuren
        light_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
            if row_idx % 2 == 0:
                for cell in row:
                    if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                        cell.fill = light_fill

        # Zet rijen vast (freeze panes)
        ws.freeze_panes = "A2"

    print(f"Excel opgeslagen: {path}")


# ─── Hoofdprogramma ───────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Scrape contractdata voor top-scoring tenders")
    parser.add_argument("--top", type=int, default=100, help="Aantal top tenders (standaard 100)")
    parser.add_argument("--min-score", type=int, default=50, help="Minimum AI score (standaard 50)")
    parser.add_argument("--delay", type=float, default=1.5, help="Wachttijd tussen requests in seconden")
    parser.add_argument("--no-scrape", action="store_true", help="Alleen dataset + cache, geen scraping")
    args = parser.parse_args()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # 1. Dataset laden
    dataset_path = find_dataset()
    if not dataset_path:
        print("FOUT: Geen TenderNed dataset gevonden.")
        sys.exit(1)

    df = load_dataset(dataset_path)
    if df is None:
        sys.exit(1)

    # 2. AI scores koppelen
    print("\nAI scores laden...")
    ai_cache = load_ai_cache()
    print(f"  {len(ai_cache):,} entries in AI cache")
    df = merge_ai_scores(df, ai_cache)

    # 3. Filteren op top tenders
    scored = df[df["ai_score"].notna()].copy()
    scored["ai_score"] = pd.to_numeric(scored["ai_score"], errors="coerce")
    top = scored[scored["ai_score"] >= args.min_score].nlargest(args.top, "ai_score").copy()

    print(f"\nTop {len(top)} tenders geselecteerd (min score: {args.min_score})")

    # 4. Scrape cache laden
    scrape_cache = load_scrape_cache()

    # 5. Per tender: contract_end bepalen via Playwright
    scraped_results = []
    needs_scraping = 0
    already_known = 0
    from_cache = 0

    # Bepaal welke URLs gescraped moeten worden
    to_scrape = []
    for _, row in top.iterrows():
        tender_url = str(row.get("tender_url", "") or "")
        contract_end_known = row.get("contract_end")
        end_known = (
            contract_end_known is not None
            and str(contract_end_known) not in ("", "nan", "NaT", "None")
        )
        to_scrape.append((tender_url, end_known))

    needs_playwright = any(
        not end_known and url and not args.no_scrape and url not in scrape_cache
        for url, end_known in to_scrape
    )

    # Start Playwright één keer voor alle requests
    playwright_ctx = None
    browser = None
    if needs_playwright:
        try:
            from playwright.sync_api import sync_playwright
            playwright_ctx = sync_playwright().start()
            browser = playwright_ctx.chromium.launch(headless=True)
            print("  Playwright browser gestart")
        except ImportError:
            print("  WAARSCHUWING: Playwright niet geinstalleerd. Installeer met: pip install playwright && playwright install chromium")

    try:
        for tender_url, end_known in to_scrape:
            empty_entry = {
                "scrape_success": False,
                "scraped_contract_start": None,
                "scraped_contract_end": None,
                "scraped_contract_duration": None,
                "scraped_contract_value": None,
                "scraped_num_bids": None,
                "scraped_url": tender_url,
                "scraped_at": None,
            }

            if end_known:
                already_known += 1
                scraped_results.append(empty_entry)
            elif not tender_url or args.no_scrape:
                scraped_results.append(empty_entry)
            elif tender_url in scrape_cache:
                from_cache += 1
                scraped_results.append(scrape_cache[tender_url])
            elif browser:
                needs_scraping += 1
                print(f"  [{needs_scraping}] Scrapen: {tender_url[-40:]}...")
                entry = scrape_tenderned_page(tender_url, playwright_browser=browser)
                scrape_cache[tender_url] = entry
                save_scrape_cache(scrape_cache)
                scraped_results.append(entry)
                time.sleep(args.delay)
            else:
                scraped_results.append(empty_entry)
    finally:
        if browser:
            browser.close()
        if playwright_ctx:
            playwright_ctx.stop()

    print(f"\nContract_end al bekend in bronbestand: {already_known}")
    print(f"Uit scrape cache: {from_cache}")
    print(f"Nieuw gescraped van TenderNed: {needs_scraping}")

    # 6. Scraped data samenvoegen met top df
    scraped_df = pd.DataFrame(scraped_results)
    top = top.reset_index(drop=True)
    scraped_df = scraped_df.reset_index(drop=True)
    top = pd.concat([top, scraped_df], axis=1)

    # 7. Beste beschikbare contract_end bepalen (met schatting als fallback)
    from src.config import DEFAULT_CONTRACT_YEARS

    def best_contract_end(row) -> tuple[str | None, str]:
        """
        Geeft (einddatum, bron) terug. Bron geeft aan hoe betrouwbaar de datum is:
        - "Bronbestand"  → exact, uit TenderNed dataset
        - "Gescraped"    → exact, van TenderNed pagina
        - "Schatting"    → berekend: gunningsdatum + 3 jaar (SkillsTown norm)
        - "Onbekend"     → geen basis voor schatting
        """
        def is_valid(val):
            return val and str(val) not in ("", "nan", "NaT", "None")

        # 1. Bronbestand (meest betrouwbaar)
        if is_valid(row.get("contract_end")):
            return str(row["contract_end"])[:10], "Bronbestand"

        # 2. Gescraped van TenderNed
        if is_valid(row.get("scraped_contract_end")):
            return str(row["scraped_contract_end"])[:10], "Gescraped"

        # 3. Schatting: gunningsdatum + DEFAULT_CONTRACT_YEARS jaar
        if is_valid(row.get("award_date")):
            try:
                end_dt = pd.to_datetime(row["award_date"]) + pd.DateOffset(years=DEFAULT_CONTRACT_YEARS)
                return end_dt.strftime("%Y-%m-%d"), f"Schatting ({DEFAULT_CONTRACT_YEARS}j na gunning)"
            except Exception:
                pass

        # 4. Schatting: publicatiedatum + 6 maanden (procurement) + DEFAULT_CONTRACT_YEARS jaar
        if is_valid(row.get("publication_date")):
            try:
                end_dt = (
                    pd.to_datetime(row["publication_date"])
                    + pd.DateOffset(months=6)
                    + pd.DateOffset(years=DEFAULT_CONTRACT_YEARS)
                )
                return end_dt.strftime("%Y-%m-%d"), f"Schatting ({DEFAULT_CONTRACT_YEARS}j na publicatie)"
            except Exception:
                pass

        return None, "Onbekend"

    results = top.apply(best_contract_end, axis=1)
    top["contract_end_final"] = results.apply(lambda x: x[0])
    top["contract_end_bron"] = results.apply(lambda x: x[1])

    # 8. Actiedatum + urgentie berekenen
    top["sales_action_date"] = top["contract_end_final"].apply(
        lambda d: calculate_action_date(d, DEFAULT_LEAD_MONTHS)
    )
    top["days_until_action"] = top["sales_action_date"].apply(days_until_action)
    top["urgency"] = top["days_until_action"].apply(urgency_label)

    # Sorteer: urgente bovenaan, daarna op AI score
    urgency_order = {"VERLOPEN": 0, "URGENT": 1, "Hoog": 2, "Medium": 3, "Laag": 4, "Onbekend": 5}
    top["urgency_sort"] = top["urgency"].map(urgency_order)
    top = top.sort_values(["urgency_sort", "ai_score"], ascending=[True, False]).drop(columns=["urgency_sort"])

    # 9. Samenvatting printen
    print(f"\n{'─'*55}")
    print(f"SAMENVATTING")
    print(f"{'─'*55}")
    print(f"Totaal tenders: {len(top)}")
    print(f"Met bekende actiedatum: {top['sales_action_date'].notna().sum()}")
    for label in ["VERLOPEN", "URGENT", "Hoog", "Medium", "Laag", "Onbekend"]:
        count = (top["urgency"] == label).sum()
        if count:
            print(f"  {label}: {count}")

    # Top 10 printen
    print(f"\nTop 10 tenders:")
    for _, row in top.head(10).iterrows():
        score = row.get("ai_score", "?")
        title = str(row.get("title", ""))[:50]
        org = str(row.get("organization", ""))[:30]
        action = row.get("sales_action_date", "onbekend")
        urgency = row.get("urgency", "?")
        print(f"  [{score:>3}] [{urgency:<8}] {title:<50} | {org:<30} | Actie: {action}")

    # 10. Exporteer naar Excel
    today_str = datetime.now().strftime("%Y-%m-%d")
    excel_path = OUTPUT_DIR / f"top_tenders_sales_{today_str}.xlsx"
    export_to_excel(top, excel_path)

    print(f"\nKlaar! Deel dit bestand met Sales: {excel_path}")


if __name__ == "__main__":
    main()
